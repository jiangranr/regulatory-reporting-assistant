"""把字段级 G31 修订对照表导出为 Excel，供主流程上传测试。

输出文件：backend/data/G31_修订对照表_v2_2026Q1.xlsx
列结构与 v2 字段级模板一致（17 列 + affected_cells 列）。
JSON 字段（before_value / after_value / 枚举值列表）按可读的多行文本展示。

affected_cells 通过查 reg_reporting_items 表按 field_code 关键字反向匹配填入。
DB 不可用时该列留空，parser 会回退到 row+col slug 匹配（向下兼容）。
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session

from app.core.database import engine
from scripts.seed_g31_revisions_v2 import REGIME, REVISIONS, expand_cells_for_field


# 列顺序 = 用户读表格的自然顺序：身份 → 变更内容 → 证据 → 衍生 → 审核
COLUMNS = [
    ("revision_id", "变更ID", 30),
    ("regime_version", "制度版本", 22),
    ("reporting_system_code", "报送体系", 12),
    ("report_code", "报表编号", 12),
    ("section_code", "分区", 10),
    ("field_code", "字段编码", 28),
    ("field_name", "业务字段名", 32),
    ("change_type", "变更类型", 12),
    ("change_dimensions", "变更维度", 22),
    ("change_summary", "变更摘要（一句话）", 55),
    ("before_value", "旧值（JSON）", 42),
    ("after_value", "新值（JSON）", 42),
    ("source_document_id", "关联发文 ID", 14),
    ("regulation_evidence", "监管原文", 55),
    ("evidence_source_ref", "原文锚点", 30),
    ("affected_cells", "影响 cell 列表", 50),
    ("effective_date", "生效日期", 14),
    ("confidence_level", "置信度", 10),
    ("review_status", "复核状态", 12),
]


def _fmt_value(field: str, val) -> str:
    if val is None:
        return ""
    if field in {"change_dimensions"} and isinstance(val, list):
        return "、".join(val)
    if field == "affected_cells" and isinstance(val, list):
        return "\n".join(val) if val else ""
    if field in {"before_value", "after_value"} and isinstance(val, dict):
        if not val:
            return "（空 / 新增字段）" if field == "before_value" else ""
        lines = []
        for k, v in val.items():
            if isinstance(v, list):
                v_str = "、".join(map(str, v))
            else:
                v_str = str(v)
            lines.append(f"{k}：{v_str}")
        return "\n".join(lines)
    return str(val)


def _build_affected_cells_map() -> dict[str, list[str]]:
    """从 DB 反查每个 field_code 对应的 affected_cells。

    DB 不可用时返回空 dict，所有 cell 留空。Parser 会回退到 row+col 匹配。
    """
    try:
        with Session(engine) as session:
            return {
                entry["field_code"]: expand_cells_for_field(
                    session, entry["report_code"], entry["field_code"]
                )
                for entry in REVISIONS
            }
    except Exception as exc:
        print(f"  ⚠ DB 不可用，affected_cells 列将留空：{exc}")
        return {}


def build_workbook() -> Workbook:
    wb = Workbook()

    # ── Sheet 1: 制度版本（regime_versions） ──────────────────────
    ws_regime = wb.active
    ws_regime.title = "制度版本"
    regime_cols = [
        ("制度版本", REGIME["regime_version"]),
        ("报送体系", REGIME["reporting_system_code"]),
        ("版本名称", REGIME["version_name"]),
        ("生效日期", REGIME["effective_date"]),
        ("发文日期", REGIME["publish_date"]),
        ("源 zip 文件", REGIME["source_zip_name"]),
        ("影响报表", "、".join(REGIME["affected_report_codes"])),
        ("修订条目数", str(len(REVISIONS))),
        ("状态", REGIME["status"]),
        ("备注", REGIME["notes"]),
    ]
    title_font = Font(bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="C8643B")  # 项目橙
    ws_regime["A1"] = "字段"
    ws_regime["B1"] = "值"
    for cell in (ws_regime["A1"], ws_regime["B1"]):
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")
    for i, (k, v) in enumerate(regime_cols, start=2):
        ws_regime.cell(i, 1, k).font = Font(bold=True)
        ws_regime.cell(i, 2, v).alignment = Alignment(wrap_text=True, vertical="top")
    ws_regime.column_dimensions["A"].width = 18
    ws_regime.column_dimensions["B"].width = 60

    # ── Sheet 2: 修订条目（regulatory_revisions） ──────────────
    ws = wb.create_sheet("修订对照表")
    # 表头
    header_fill = PatternFill("solid", fgColor="F4A56B")
    for col_idx, (_, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, col_idx, label)
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # 数据
    change_type_colors = {
        "NEW": "C6EFCE",        # 浅绿
        "MODIFY": "FFEB9C",     # 浅黄
        "DELETE": "FFC7CE",     # 浅红
    }
    cells_map = _build_affected_cells_map()
    for row_idx, entry in enumerate(REVISIONS, start=2):
        # 准备一份完整 entry（含 regime 默认字段）
        full_entry = {
            **entry,
            "regime_version": REGIME["regime_version"],
            "reporting_system_code": REGIME["reporting_system_code"],
            "source_document_id": None,
            "effective_date": REGIME["effective_date"],
            "review_status": "DRAFT",
            "affected_cells": cells_map.get(entry["field_code"], []),
        }
        ct_fill = PatternFill("solid", fgColor=change_type_colors.get(entry["change_type"], "FFFFFF"))
        for col_idx, (field, _, _) in enumerate(COLUMNS, start=1):
            v = full_entry.get(field)
            cell = ws.cell(row_idx, col_idx, _fmt_value(field, v))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if field == "change_type":
                cell.fill = ct_fill
                cell.font = Font(bold=True)
        ws.row_dimensions[row_idx].height = 90

    # ── Sheet 3: 字段说明（READ ME） ──────────────────────────
    ws_help = wb.create_sheet("字段说明")
    ws_help["A1"] = "列名"
    ws_help["B1"] = "类型"
    ws_help["C1"] = "说明"
    for cell in (ws_help["A1"], ws_help["B1"], ws_help["C1"]):
        cell.font = title_font
        cell.fill = title_fill
    descriptions = [
        ("revision_id", "STRING", "全局唯一编号。格式建议：REV-{system}-{period}-{report}-{seq}"),
        ("regime_version", "STRING", "制度版本号，关联制度版本 sheet。同一字段同一版本只能有一条记录。"),
        ("reporting_system_code", "STRING", "1104 / EAST / 一表通 / 客户风险"),
        ("report_code", "STRING", "受影响报表代码，如 G31。"),
        ("section_code", "STRING", "报表分区，如 PART_I / MAIN。"),
        ("field_code", "STRING", "字段技术编码（snake_case 英文），系统级唯一。"),
        ("field_name", "STRING", "业务字段中文名。"),
        ("change_type", "ENUM", "NEW / MODIFY / DELETE / RENAME / SPLIT / MERGE"),
        ("change_dimensions", "ENUM[]", "14 项之一或多选：口径/数据类型/长度/单位/枚举值/必填/默认值/校验规则/计算公式/填报说明/归集范围/数据来源系统/报送频率/机构范围"),
        ("change_summary", "TEXT", "**一句话变更摘要（人话）**。列表展示用，禁止只写「口径变化」之类抽象描述，必须说清「具体怎么变」。"),
        ("before_value", "JSON-OBJECT", "旧版本属性快照。新增字段为空。Excel 里用「属性：值」多行展示。"),
        ("after_value", "JSON-OBJECT", "新版本属性快照。删除字段为空。Excel 里用「属性：值」多行展示。"),
        ("source_document_id", "INT (FK)", "关联 reg_documents.id，未来 zip 解析填，手工填可留空。"),
        ("regulation_evidence", "TEXT", "监管原文片段，证明这条修订的依据。"),
        ("evidence_source_ref", "STRING", "文件锚点：文件名#章节号，未来用于跳转原文高亮。"),
        ("affected_cells", "ITEM_CODE[]", "该字段映射到的具体报表 cell 列表（如 G31.PART_I.1_0.D_...）。多个用换行/逗号分隔。留空时主流程会按 row+col 兜底匹配。"),
        ("effective_date", "DATE", "首次生效报送期，未填则默认取制度版本的生效日期。"),
        ("confidence_level", "ENUM", "AI 抽取置信度：HIGH / MEDIUM / LOW。"),
        ("review_status", "ENUM", "DRAFT（草稿）/ CONFIRMED（已确认）/ DISPUTED（有争议）"),
    ]
    for i, (k, t, desc) in enumerate(descriptions, start=2):
        ws_help.cell(i, 1, k).font = Font(bold=True, name="Menlo")
        ws_help.cell(i, 2, t).font = Font(name="Menlo")
        ws_help.cell(i, 3, desc).alignment = Alignment(wrap_text=True, vertical="top")
        ws_help.row_dimensions[i].height = 36
    ws_help.column_dimensions["A"].width = 24
    ws_help.column_dimensions["B"].width = 14
    ws_help.column_dimensions["C"].width = 80

    return wb


def main() -> Path:
    out_dir = Path(__file__).resolve().parent.parent / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "G31_修订对照表_v2_2026Q1.xlsx"
    wb = build_workbook()
    wb.save(out_path)
    print(f"✓ 已生成 {out_path}")
    print(f"  共 3 个 sheet：制度版本 / 修订对照表（{len(REVISIONS)} 条）/ 字段说明")
    print(f"  生成时间：{datetime.now().isoformat()[:19]}")
    return out_path


if __name__ == "__main__":
    main()
