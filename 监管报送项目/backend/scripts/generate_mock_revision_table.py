"""
生成 G31(251) 修订对照表 mock 文件，用于测试 revision_table_parser + item_change_scanner。

基于已知的 G31 251 版变更信息：
  列 C（修正久期）: NEW      —— 2025年新增
  列 D（间接持有）: MODIFIED —— 穿透统计口径调整
  列 E（穿透后余额）: MODIFIED
  列 A（穿透前余额）: UNCHANGED
  列 B（投资收入）:  UNCHANGED

输出文件：backend/data/G31_修订对照表_mock.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

# 确保 openpyxl 可用
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("请先安装 openpyxl：pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent.parent
OUT_PATH = ROOT / "data" / "G31_修订对照表_mock.xlsx"
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# ── 基于 G31(251).xls 已知的行列结构 ───────────────────────────────────────────
# 行标签（代表性行，不完整，仅供测试）
ROW_LABELS = [
    "债券投资合计",
    "一、国债",
    "二、地方政府债",
    "三、政策性银行债",
    "四、商业银行债",
    "五、企业债",
    "六、可转换公司债",
    "七、其他债券",
    "股权及其他投资合计",
    "一、股权投资",
    "二、基金投资合计",
    "（一）货币市场基金",
    "（二）债券型基金",
    "（三）混合型基金",
    "（四）股票型基金",
    "（五）其他基金",
    "三、其他投资",
]

# 列标签 → (change_type, change_desc)
COL_RULES = {
    "A_穿透前_期末余额": ("UNCHANGED", "原有指标，口径无变化"),
    "B_投资收入_年初至报告期末数": ("UNCHANGED", "原有指标，口径无变化"),
    "C_修正久期": ("NEW", "2025年251版新增，用于量化债券市场价值波动风险，仅适用于债券类投资"),
    "D_因持有非底层资产而间接持有": ("MODIFIED", "251版调整穿透统计口径，需重新梳理数据源"),
    "穿透后_期末余额": ("MODIFIED", "251版调整穿透统计口径，与D列联动调整"),
}

HEADERS = ["表号", "分区", "行标识", "列标识", "变更类型", "变更说明"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "G31修订对照表"

# 写表头
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# 写数据行
row_idx = 2
for row_label in ROW_LABELS:
    for col_suffix, (change_type, desc) in COL_RULES.items():
        # C 列（修正久期）仅适用于债券类，股权/基金类跳过
        if "修正久期" in col_suffix and "债券" not in row_label and "国债" not in row_label \
                and "地方政府" not in row_label and "政策性" not in row_label \
                and "商业银行债" not in row_label and "企业债" not in row_label \
                and "可转换" not in row_label and "其他债券" not in row_label:
            continue

        ws.cell(row=row_idx, column=1, value="G31")
        ws.cell(row=row_idx, column=2, value="PART_I")
        ws.cell(row=row_idx, column=3, value=row_label)
        ws.cell(row=row_idx, column=4, value=col_suffix)
        ws.cell(row=row_idx, column=5, value=change_type)
        ws.cell(row=row_idx, column=6, value=desc)

        # 不同变更类型不同背景色
        fill_color = {
            "NEW": "C6EFCE",
            "MODIFIED": "FFEB9C",
            "DELETED": "FFC7CE",
            "UNCHANGED": "FFFFFF",
        }.get(change_type, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for c in range(1, 7):
            ws.cell(row=row_idx, column=c).fill = row_fill

        row_idx += 1

# 调整列宽
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 30
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 40

wb.save(OUT_PATH)
print(f"Mock 修订对照表已生成：{OUT_PATH}")
print(f"  共 {row_idx - 2} 行变更记录")
