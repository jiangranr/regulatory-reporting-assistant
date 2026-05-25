"""
生成更全面的 G31 修订对照表测试 fixture。

相对 `generate_mock_revision_table.py`：
- 覆盖 PART_I（投资业务情况）/ PART_II（底层资产穿透）/ PART_III（集中度）三个分区
- 同时覆盖行变更与列变更
- 四种 change_type（NEW / MODIFIED / DELETED / UNCHANGED）都有代表
- 行/列命名贴近 1104 G31 251 版实际口径

修订内容参考公开信息（1104 G31 报表历次修订惯例）：
  - 修正久期：新增列，量化债券市场价值波动风险，仅适用债券类
  - 穿透后口径：调整，与"间接持有"列联动
  - 资产支持证券、同业存单：细分行新增
  - 标准化债权类资产：口径调整
  - 非标-应收账款：废止细项
  - 集中度风险加权权重：新增列

输出路径：backend/tests/fixtures/G31_修订对照表.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl：pip install openpyxl")
    sys.exit(1)


ROOT = Path(__file__).parent.parent
OUT_PATH = ROOT / "tests" / "fixtures" / "G31_修订对照表.xlsx"
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)


HEADERS = ["表号", "分区", "行标识", "列标识", "变更类型", "变更说明"]


# ── 行配置 ─────────────────────────────────────────────────────────────────────
# (row_label, row_change_type, row_desc)
# row_change_type 仅在该行整体新增/删除时填 NEW/DELETED；
# 其他情况下放 UNCHANGED，由列变更承担细节
PART_I_ROWS = [
    ("债券投资合计", "UNCHANGED", ""),
    ("一、国债", "UNCHANGED", ""),
    ("二、地方政府债", "UNCHANGED", ""),
    ("三、政策性银行债", "UNCHANGED", ""),
    ("四、商业银行债", "UNCHANGED", ""),
    ("五、企业债", "UNCHANGED", ""),
    ("六、可转换公司债", "UNCHANGED", ""),
    ("七、同业存单", "NEW", "251 版从'其他债券'中独立出来作为单独细分行，便于穿透同业敞口"),
    ("八、资产支持证券（ABS/MBS）", "NEW", "251 版新增，区分标准化与非标资产证券化产品"),
    ("九、其他债券", "MODIFIED", "口径调整：剔除同业存单和资产支持证券，仅保留剩余债券"),
    ("股权及其他投资合计", "UNCHANGED", ""),
    ("一、股权投资", "UNCHANGED", ""),
    ("二、基金投资合计", "MODIFIED", "251 版要求按底层资产类别拆分基金，原仅按基金类型分类"),
    ("（一）货币市场基金", "UNCHANGED", ""),
    ("（二）债券型基金", "UNCHANGED", ""),
    ("（三）混合型基金", "UNCHANGED", ""),
    ("（四）股票型基金", "UNCHANGED", ""),
    ("（五）FOF/MOM 基金", "NEW", "251 版新增，FOF 基金需穿透至底层组合"),
    ("（六）其他基金", "UNCHANGED", ""),
    ("三、信托及资管计划", "MODIFIED", "口径调整：纳入券商资管、保险资管，原仅含信托"),
    ("四、应收账款类投资", "DELETED", "251 版废止该细项，已统一归入'其他投资'"),
    ("五、其他投资", "MODIFIED", "口径调整：吸收已废止的'应收账款类投资'"),
]

PART_II_ROWS = [
    ("底层资产合计", "UNCHANGED", ""),
    ("一、标准化债权类资产", "MODIFIED", "251 版口径细化：明确债券、ABS、同业存单的归属"),
    ("二、非标准化债权类资产", "MODIFIED", "排除标债定义后剩余口径"),
    ("（一）信托贷款", "UNCHANGED", ""),
    ("（二）委托贷款", "UNCHANGED", ""),
    ("（三）应收账款", "DELETED", "251 版废止，相关数据并入其他非标"),
    ("（四）票据资产", "UNCHANGED", ""),
    ("（五）其他非标", "MODIFIED", "口径调整：吸收已废止的'应收账款'细项"),
    ("三、权益类资产", "UNCHANGED", ""),
    ("四、商品及其他", "UNCHANGED", ""),
]

PART_III_ROWS = [
    ("一、最大单一发行人集中度", "UNCHANGED", ""),
    ("二、最大十家发行人集中度", "NEW", "251 版新增，监测发行人集中度风险"),
    ("三、同一行业集中度", "UNCHANGED", ""),
    ("四、同一区域集中度", "UNCHANGED", ""),
]


# ── 列配置（按分区分别声明） ───────────────────────────────────────────────────
# (col_label, change_type, change_desc, [applicable_row_predicate])
# applicable_row_predicate 用于实现"列变更只在部分行生效"（例如修正久期仅适用债券）

def _is_bond_row(row_label: str) -> bool:
    keywords = ("国债", "地方政府债", "政策性银行债", "商业银行债",
                "企业债", "可转换", "同业存单", "资产支持证券", "其他债券", "债券投资合计")
    return any(k in row_label for k in keywords)


PART_I_COLUMNS = [
    ("A_穿透前_期末余额", "UNCHANGED", "原有指标，口径无变化", None),
    ("B_投资收入_年初至报告期末数", "UNCHANGED", "原有指标，口径无变化", None),
    ("C_修正久期", "NEW", "251 版新增，量化债券市场价值波动风险，仅适用债券类投资", _is_bond_row),
    ("D_因持有非底层资产而间接持有", "MODIFIED", "251 版调整穿透统计口径，需重新梳理数据源", None),
    ("E_穿透后_期末余额", "MODIFIED", "251 版调整穿透统计口径，与 D 列联动调整", None),
]

PART_II_COLUMNS = [
    ("A_名义本金", "UNCHANGED", "原有指标，口径无变化", None),
    ("B_账面价值", "UNCHANGED", "原有指标，口径无变化", None),
    ("C_市值/公允价值", "MODIFIED", "251 版要求统一采用公允价值口径，原可按账面价值估计", None),
    ("D_减值准备", "UNCHANGED", "原有指标，口径无变化", None),
    ("E_穿透层级", "NEW", "251 版新增，记录从合并报表到底层资产的穿透层数", None),
]

PART_III_COLUMNS = [
    ("A_余额", "UNCHANGED", "原有指标，口径无变化", None),
    ("B_占比", "UNCHANGED", "原有指标，口径无变化", None),
    ("C_风险加权资产权重", "NEW", "251 版新增，反映集中度风险加权计量", None),
]


SECTION_ROWS = {
    "PART_I": PART_I_ROWS,
    "PART_II": PART_II_ROWS,
    "PART_III": PART_III_ROWS,
}
SECTION_COLUMNS = {
    "PART_I": PART_I_COLUMNS,
    "PART_II": PART_II_COLUMNS,
    "PART_III": PART_III_COLUMNS,
}


# ── 写 Excel ───────────────────────────────────────────────────────────────────

CHANGE_TYPE_FILL = {
    "NEW": "C6EFCE",
    "MODIFIED": "FFEB9C",
    "DELETED": "FFC7CE",
    "UNCHANGED": "FFFFFF",
}


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "G31修订对照表"

    # 表头
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    stats = {"NEW": 0, "MODIFIED": 0, "DELETED": 0, "UNCHANGED": 0}

    for section, rows in SECTION_ROWS.items():
        cols = SECTION_COLUMNS[section]
        for row_label, row_change, row_desc in rows:
            # 行整体变更：行新增/删除时输出一条"行级"记录，列标识用 "*"
            if row_change in {"NEW", "DELETED"}:
                _write_row(ws, row_idx, section, row_label, "*", row_change, row_desc)
                stats[row_change] += 1
                row_idx += 1

            for col_label, col_change, col_desc, predicate in cols:
                if predicate is not None and not predicate(row_label):
                    continue
                # 已删除行不再展开列
                if row_change == "DELETED":
                    continue
                # 整行新增 + 列原本不变 → 跳过列级 UNCHANGED 噪音
                if row_change == "NEW" and col_change == "UNCHANGED":
                    continue

                _write_row(ws, row_idx, section, row_label, col_label, col_change, col_desc)
                stats[col_change] += 1
                row_idx += 1

    # 列宽
    widths = {"A": 8, "B": 10, "C": 32, "D": 32, "E": 12, "F": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(OUT_PATH)

    total = row_idx - 2
    print(f"已生成 G31 修订对照表：{OUT_PATH}")
    print(f"  共 {total} 条变更记录")
    for change_type in ("NEW", "MODIFIED", "DELETED", "UNCHANGED"):
        print(f"    {change_type:10}: {stats[change_type]} 条")


def _write_row(ws, row_idx: int, section: str, row_label: str,
               col_label: str, change_type: str, desc: str) -> None:
    ws.cell(row=row_idx, column=1, value="G31")
    ws.cell(row=row_idx, column=2, value=section)
    ws.cell(row=row_idx, column=3, value=row_label)
    ws.cell(row=row_idx, column=4, value=col_label)
    ws.cell(row=row_idx, column=5, value=change_type)
    ws.cell(row=row_idx, column=6, value=desc)

    fill_color = CHANGE_TYPE_FILL.get(change_type, "FFFFFF")
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for c in range(1, 7):
        ws.cell(row=row_idx, column=c).fill = fill


if __name__ == "__main__":
    main()
