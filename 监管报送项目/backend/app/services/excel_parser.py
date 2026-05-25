"""Excel 表样解析器：从 .xls/.xlsx 提取 template_cells / dimension_members / reporting_items。"""
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

import xlrd


# xlrd 颜色索引映射（基于 G31(251).xls 实测）
_DERIVED_COLOR_INDICES = {46}   # 紫色/lavender → DERIVED
_NO_DATA_COLOR_INDICES = {22}   # 灰色 → NO_DATA
_FILLABLE_COLOR_INDICES = {9}   # 纯白（显式） → FILLABLE
# color=64 为 auto（默认无填充），根据位置判断


@dataclass
class ExcelParseResult:
    template_cells: list[dict] = field(default_factory=list)
    dimension_members: list[dict] = field(default_factory=list)
    items: list[dict] = field(default_factory=list)
    item_dimensions: list[dict] = field(default_factory=list)


@dataclass(frozen=True)
class ExcelDataLayout:
    data_row_start: int
    data_cols: list[int]
    header_row_start: int


def _col_letter(col_idx: int) -> str:
    """0-based 列索引 → Excel 列字母。0→A, 3→D"""
    result = ""
    n = col_idx + 1
    while n:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _excel_ref(row: int, col: int) -> str:
    """0-based (row, col) → Excel 引用。(7, 3) → 'D8'"""
    return f"{_col_letter(col)}{row + 1}"


def _slugify(text: str) -> str:
    """将中文标签转为 item_code 片段。"""
    text = re.sub(r'[^\w一-鿿]', '_', text.strip())
    text = re.sub(r'_+', '_', text).strip('_')
    return text[:40] if text else "UNKNOWN"


def _cell_role_from_color(color_index: int, row: int, data_row_start: int,
                          col: int, data_col_start: int) -> str:
    """根据颜色索引和位置推断单元格角色。"""
    if color_index in _NO_DATA_COLOR_INDICES:
        return "NO_DATA"
    if color_index in _DERIVED_COLOR_INDICES:
        return "DERIVED"
    if color_index in _FILLABLE_COLOR_INDICES:
        return "FILLABLE"
    # color=64（auto）：在数据区域内视为 FILLABLE，否则 HEADER
    if row >= data_row_start and col >= data_col_start:
        return "FILLABLE"
    return "HEADER"


def _cell_role_from_layout(
    color_index: int,
    row: int,
    col: int,
    layout: ExcelDataLayout,
) -> str:
    if color_index in _NO_DATA_COLOR_INDICES:
        return "NO_DATA"
    if color_index in _DERIVED_COLOR_INDICES:
        return "DERIVED"
    if color_index in _FILLABLE_COLOR_INDICES:
        return "FILLABLE"
    if row >= layout.data_row_start and col in layout.data_cols:
        return "FILLABLE"
    return "HEADER"


def _detect_data_boundary(sheet: xlrd.sheet.Sheet) -> tuple[int, int]:
    """
    自动检测数据区域起始行和起始列。
    启发式：找到第一行有数字值（ctype=2）的行，以及包含数字的最小列。
    """
    for r in range(min(sheet.nrows, 20)):
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_NUMBER:
                return r, c
    import warnings
    warnings.warn(
        f"Could not auto-detect data boundary in sheet '{sheet.name}', defaulting to (0, 0). "
        "Consider providing explicit boundaries.",
        UserWarning,
        stacklevel=2,
    )
    return 0, 0


def _detect_data_layout(sheet: xlrd.sheet.Sheet) -> ExcelDataLayout:
    """Detect business data rows and reporting measure columns.

    G31-style 1104 templates have row numbers/business labels on the left and
    regulatory measure columns marked by header letters such as A/B/C/D/E.
    The previous boundary detector returned column A because the first data
    row contains the row number "1.0", which polluted diff results with
    structural columns. This layout detector anchors on the regulatory header
    letters instead.
    """
    data_row_start, fallback_col_start = _detect_data_boundary(sheet)
    header_row, data_cols = _detect_measure_header_columns(sheet, data_row_start)
    if not data_cols:
        data_cols = list(range(fallback_col_start, sheet.ncols))
        header_row = max(0, data_row_start - 1)
    return ExcelDataLayout(
        data_row_start=data_row_start,
        data_cols=data_cols,
        header_row_start=header_row,
    )


def _detect_measure_header_columns(
    sheet: xlrd.sheet.Sheet,
    data_row_start: int,
) -> tuple[int, list[int]]:
    best_row = 0
    best_cols: list[int] = []
    for row in range(data_row_start):
        letter_cols = [
            col
            for col in range(sheet.ncols)
            if re.fullmatch(r"[A-Z]{1,2}", str(sheet.cell_value(row, col)).strip())
        ]
        for group in _consecutive_groups(letter_cols):
            if len(group) > len(best_cols):
                best_row = row
                best_cols = group
    return best_row, best_cols


def _consecutive_groups(values: list[int]) -> list[list[int]]:
    if not values:
        return []
    groups: list[list[int]] = []
    current = [values[0]]
    for value in values[1:]:
        if value == current[-1] + 1:
            current.append(value)
        else:
            groups.append(current)
            current = [value]
    groups.append(current)
    return groups


def _business_row_label(sheet: xlrd.sheet.Sheet, row: int, data_col_start: int) -> str:
    for col in range(data_col_start - 1, -1, -1):
        value = str(sheet.cell_value(row, col)).strip()
        if value and not re.fullmatch(r"\d+(?:\.0+)?", value):
            return value
    for col in range(data_col_start - 1, -1, -1):
        value = str(sheet.cell_value(row, col)).strip()
        if value:
            return value
    return f"ROW_{row}"


def _is_non_business_row_label(label: str) -> bool:
    return any(keyword in label for keyword in ("填表人", "复核人", "负责人", "联系人", "联系电话"))


def parse_excel(
    file_path: Path,
    object_code: str,
    section_code: str = "PART_I",
) -> ExcelParseResult:
    """
    解析 .xls 文件，返回 ExcelParseResult。
    支持 .xls（xlrd）和 .xlsx（openpyxl）。
    """
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx(file_path, object_code, section_code)

    # .xls 使用 xlrd（需要 formatting_info=True 读取颜色）
    try:
        wb = xlrd.open_workbook(str(file_path), formatting_info=True)
    except Exception as exc:
        raise ValueError(f"无法打开 Excel 文件 {file_path}: {exc}") from exc
    sheet = wb.sheet_by_index(0)
    sheet_name = sheet.name

    result = ExcelParseResult()
    _tc_role_map: dict[tuple[int, int], str] = {}
    layout = _detect_data_layout(sheet)
    data_col_start = min(layout.data_cols) if layout.data_cols else 0

    # ── 1. 提取所有 template_cells ──────────────────────────────────────────
    merge_map: dict[tuple[int, int], dict] = {}
    for r1, r2, c1, c2 in sheet.merged_cells:
        for r in range(r1, r2):
            for c in range(c1, c2):
                merge_map[(r, c)] = {"r1": r1, "r2": r2, "c1": c1, "c2": c2}

    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            cell = sheet.cell(row, col)
            try:
                xfi = sheet.cell_xf_index(row, col)
                xf = wb.xf_list[xfi]
                color_index = xf.background.pattern_colour_index
            except Exception:
                color_index = 64

            role = _cell_role_from_layout(color_index, row, col, layout)
            raw_text = str(cell.value).strip() if cell.value else ""

            style = {"color_index": color_index, "role": role}
            merge = merge_map.get((row, col), {})

            result.template_cells.append({
                "sheet_name": sheet_name,
                "row_index": row,
                "col_index": col,
                "excel_ref": _excel_ref(row, col),
                "raw_text": raw_text,
                "cell_type": role,
                "style_json": json.dumps(style),
                "merge_json": json.dumps(merge),
            })
            _tc_role_map[(row, col)] = role

    # ── 2. 提取列维度成员（度量列头：A/B/C... 行到数据区前一行）──────────────
    col_labels: dict[int, str] = {}
    for col in layout.data_cols:
        label_parts = []
        for row in range(layout.header_row_start, layout.data_row_start):
            cell = sheet.cell(row, col)
            txt = str(cell.value).strip() if cell.value else ""
            if txt:
                label_parts.append(txt)
        label = "·".join(label_parts) if label_parts else f"COL_{col}"
        col_labels[col] = label
        col_slug = _slugify(label)
        result.dimension_members.append({
            "member_code": f"{object_code}.{section_code}.COL.{col_slug}",
            "member_name": label,
            "axis": "COLUMN",
            "display_order": col - data_col_start,
            "source_cell_ref": _excel_ref(layout.data_row_start - 1, col),
        })

    # ── 3. 提取行维度成员（行标签：data_col_start-1 列，data_row_start 以下）──
    row_labels: dict[int, str] = {}
    label_col = max(data_col_start - 1, 0)

    for row in range(layout.data_row_start, sheet.nrows):
        label = _business_row_label(sheet, row, data_col_start)
        row_labels[row] = label
        row_slug = _slugify(label)
        result.dimension_members.append({
            "member_code": f"{object_code}.{section_code}.ROW.{row_slug}",
            "member_name": label,
            "axis": "ROW",
            "display_order": row - layout.data_row_start,
            "source_cell_ref": _excel_ref(row, label_col),
        })

    # ── 4. 生成 reporting_items（数据区域内非 NO_DATA 单元格）────────────────
    seen_codes: set[str] = set()
    for row in range(layout.data_row_start, sheet.nrows):
        row_label = row_labels.get(row, f"ROW_{row}")
        if _is_non_business_row_label(row_label):
            continue
        row_slug = _slugify(row_label)

        for col in layout.data_cols:
            col_label = col_labels.get(col, f"COL_{col}")
            col_slug = _slugify(col_label)

            # 从已计算的 template_cells 取 role
            role = _tc_role_map.get((row, col), "NO_DATA")

            if role == "NO_DATA" or role == "HEADER":
                continue

            item_code = f"{object_code}.{section_code}.{row_slug}.{col_slug}"
            # 去重：若 item_code 已存在，加数字后缀
            original_code = item_code
            suffix_n = 1
            while item_code in seen_codes:
                item_code = f"{original_code}_{suffix_n}"
                suffix_n += 1
            seen_codes.add(item_code)
            item_name = f"{row_label}-{col_label}"

            result.items.append({
                "item_code": item_code,
                "item_name": item_name,
                "source_cell_ref": _excel_ref(row, col),
                "cell_role": role,
                "row_label": row_label,
                "column_label": col_label,
                "is_fillable": role == "FILLABLE",
                "is_derived": role == "DERIVED",
                "data_type": "DECIMAL",
                "item_type": "DERIVED" if role == "DERIVED" else "MEASURE",
            })

            result.item_dimensions.append({
                "item_code": item_code,
                "row_member_code": f"{object_code}.{section_code}.ROW.{row_slug}",
                "col_member_code": f"{object_code}.{section_code}.COL.{col_slug}",
            })

    return result


def _parse_xlsx(file_path: Path, object_code: str, section_code: str) -> ExcelParseResult:
    """解析 .xlsx 文件（用 openpyxl）。颜色检测用 RGB 值。"""
    import openpyxl

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb.active
    if ws is None:
        return ExcelParseResult()

    result = ExcelParseResult()
    sheet_name = ws.title

    def _openpyxl_role(cell) -> str:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.type == "rgb":
            rgb = fill.fgColor.rgb.upper()
            if rgb in ("FFBFBFBF", "FF808080", "FFC0C0C0", "FFD9D9D9"):
                return "NO_DATA"
            if rgb in ("FFCC99FF", "FF9966FF", "FFCCAAFF"):
                return "DERIVED"
        return "FILLABLE"

    for row in ws.iter_rows():
        for cell in row:
            r, c = cell.row - 1, cell.column - 1
            role = _openpyxl_role(cell)
            raw_text = str(cell.value).strip() if cell.value is not None else ""
            result.template_cells.append({
                "sheet_name": sheet_name,
                "row_index": r,
                "col_index": c,
                "excel_ref": _excel_ref(r, c),
                "raw_text": raw_text,
                "cell_type": role,
                "style_json": "{}",
                "merge_json": "{}",
            })

    return result
