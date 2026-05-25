"""修订对照表 Excel 解析器（v2 字段维度）

新版格式（v2 字段维度，3 sheet）：
  Sheet 1「制度版本」：版本头键值对（regime_version / effective_date / ...）
  Sheet 2「修订对照表」：每条字段级修订，17 列
  Sheet 3「字段说明」：字典说明（机器忽略）

老版格式（v1 cell 维度，单 sheet 6 列）：
  表号 | 分区 | 行标识 | 列标识 | 变更类型 | 变更说明

向下兼容：parser 自动识别两种格式。
v1 文件解析后 RevisionEntry 的新字段全部为默认空值。
v2 文件解析后 regime_version / effective_date 从 Sheet 1 注入到每条 entry。

支持 .xls（xlrd）和 .xlsx（openpyxl）。
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


# ─────────────────────────────────────────────────────────────────────
# 17 字段 RevisionEntry —— v1/v2 共用，v1 解析时新字段保持默认空。
# ─────────────────────────────────────────────────────────────────────


@dataclass
class RevisionEntry:
    # —— 原 6 字段（v1 + v2 都用）——
    table_code: str = ""          # G31
    section: str = "PART_I"       # PART_I
    row_label: str = ""           # "1.债券投资合计" / "10.0"
    column_label: str = ""        # "C·修正久期"
    change_type: str = "UNCHANGED"  # NEW / MODIFIED / DELETED / UNCHANGED
    change_desc: str = ""         # 兼容老 change_desc，新版可空（由 change_summary 替代）

    # —— v2 新增 11 字段 ——
    revision_id: str = ""                              # "REV-1104-2026Q1-G31-007"
    regime_version: str = ""                           # "1104-2026Q1-v2"
    field_code: str = ""                               # snake_case "indirect_holding_balance"
    field_name: str = ""                               # 业务字段中文名
    change_dimensions: list[str] = field(default_factory=list)   # ["口径", "校验规则"]
    change_summary: str = ""                           # 一句话人话摘要
    before_value: dict[str, Any] = field(default_factory=dict)   # JSON 反序列化
    after_value: dict[str, Any] = field(default_factory=dict)
    source_document_id: int | None = None              # 关联 reg_documents.id
    regulation_evidence: str = ""                      # 监管原文
    evidence_source_ref: str = ""                      # 文件锚点 "filing.pdf#§3.4"
    affected_cells: list[str] = field(default_factory=list)      # item_code 列表
    effective_date: str = ""                           # ISO 日期 "2026-06-30"
    confidence_level: str = "MEDIUM"                   # HIGH / MEDIUM / LOW
    review_status: str = "DRAFT"                       # DRAFT / CONFIRMED / DISPUTED


_VALID_CHANGE_TYPES = {"NEW", "MODIFIED", "DELETED", "UNCHANGED", "MODIFY", "DELETE",
                       "RENAME", "SPLIT", "MERGE"}

# 列名别名映射 —— 兼容中文表头 + 英文 snake_case + 老 v1 别名
_HEADER_ALIASES = {
    "table_code": {"表号", "报表代码", "报表", "表名", "报表编号", "table", "table_code",
                   "object_code", "report_code"},
    "section": {"分区", "部分", "区域", "sheet", "section", "section_code"},
    "row_label": {"行标识", "行名称", "行项目", "指标项目", "项目", "row", "row_label"},
    "column_label": {"列标识", "列名称", "列项目", "指标字段", "字段", "column", "column_label"},
    "change_type": {"变更类型", "调整类型", "变化类型", "类型", "change_type"},
    "change_desc": {"变更说明", "调整说明", "说明", "修订说明", "change_desc", "description"},

    # ── v2 新增 ──
    "revision_id": {"变更id", "变更编号", "revision_id", "rev_id"},
    "regime_version": {"制度版本", "版本", "regime_version", "version"},
    "reporting_system_code": {"报送体系", "体系", "reporting_system_code"},
    "field_code": {"字段编码", "字段代码", "字段标识", "field_code", "field"},
    "field_name": {"业务字段名", "字段名称", "业务字段", "field_name"},
    "change_dimensions": {"变更维度", "维度", "change_dimensions", "dimensions"},
    "change_summary": {"变更摘要", "变更摘要(一句话)", "变更摘要（一句话）", "摘要", "一句话摘要",
                       "change_summary", "summary"},
    "before_value": {"旧值", "旧值(json)", "旧值（json）", "原值", "before", "before_value"},
    "after_value": {"新值", "新值(json)", "新值（json）", "现值", "after", "after_value"},
    "source_document_id": {"关联发文id", "关联发文 id", "发文id", "source_document_id"},
    "regulation_evidence": {"监管原文", "原文", "依据原文", "regulation_evidence", "evidence"},
    "evidence_source_ref": {"原文锚点", "锚点", "源文件锚点", "evidence_source_ref"},
    "affected_cells": {"影响cell", "影响cell列表", "受影响cell", "affected_cells", "cells"},
    "effective_date": {"生效日期", "生效报送期", "effective_date"},
    "confidence_level": {"置信度", "置信", "confidence_level", "confidence"},
    "review_status": {"复核状态", "审核状态", "状态", "review_status"},
}


# ─────────────────────────────────────────────────────────────────────
# 公开 API
# ─────────────────────────────────────────────────────────────────────


def parse_revision_table(file_path: Path) -> list[RevisionEntry]:
    """解析修订对照表 Excel。自动识别 v1 / v2 格式。"""
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx(file_path)
    return _parse_xls(file_path)


# ─────────────────────────────────────────────────────────────────────
# 内部 helpers
# ─────────────────────────────────────────────────────────────────────


def _clean_header(text: str) -> str:
    return text.strip().lower().replace(" ", "").replace("_", "").replace("（", "(").replace("）", ")")


def _alias_map() -> dict[str, str]:
    return {
        _clean_header(alias): field_name
        for field_name, aliases in _HEADER_ALIASES.items()
        for alias in aliases
    }


def _normalise_change_type(raw: str) -> str:
    upper = raw.strip().upper()
    mapping = {
        "新增": "NEW",
        "ADD": "NEW",
        "修改": "MODIFIED",
        "调整": "MODIFIED",
        "MODIFY": "MODIFIED",
        "删除": "DELETED",
        "停报": "DELETED",
        "DELETE": "DELETED",
        "不变": "UNCHANGED",
        "无变化": "UNCHANGED",
    }
    if upper in _VALID_CHANGE_TYPES:
        return upper
    return mapping.get(upper, mapping.get(raw.strip(), "UNCHANGED"))


def _detect_header_map(rows: list[list[str]]) -> tuple[int, dict[str, int]] | None:
    """在前 10 行里探测表头位置，返回 (header_row_idx, {field_name: col_idx})。"""
    am = _alias_map()
    for row_idx, cells in enumerate(rows[:10]):
        mapping: dict[str, int] = {}
        for idx, cell in enumerate(cells):
            f = am.get(_clean_header(cell))
            if f:
                mapping[f] = idx
        required = {"table_code", "row_label", "column_label", "change_type"}
        # v2 表头可能没有 row_label/column_label，但有 field_code → 也算 v2
        v2_required = {"field_code", "change_type"}
        if required.issubset(mapping) or v2_required.issubset(mapping):
            return row_idx, mapping
    return None


def _cell(cells: list[str], idx: int | None) -> str:
    if idx is None or idx >= len(cells):
        return ""
    return cells[idx].strip()


# ── 多行 key:value 解析 → dict ─────────────────────────────────────


_KV_PATTERN = re.compile(r"^\s*([^：:\n]+?)\s*[：:]\s*(.+?)\s*$")


def _parse_before_after(raw: str) -> dict[str, Any]:
    """解析「key：value\nkey：value」多行格式为 dict。

    示例输入：
        口径：穿透统计含 ABS + 资管产品
        校验规则：A + D = E（不变）

    特殊处理：
      - "（空 / 新增字段）" / "(空)" / "" → 返回空 dict
      - "枚举值：境内银行、境内非银金融机构" → value 进一步切分为 list
    """
    if not raw or raw.strip() in {"（空 / 新增字段）", "(空)", "（空）", "-"}:
        return {}
    result: dict[str, Any] = {}
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        m = _KV_PATTERN.match(line)
        if not m:
            continue
        key, val = m.group(1).strip(), m.group(2).strip()
        # 枚举值 / 列表类 value 用、或,切分
        if key in {"枚举值", "归集范围", "数据来源系统"} and ("、" in val or "," in val):
            result[key] = [v.strip() for v in re.split(r"[、,]", val) if v.strip()]
        else:
            result[key] = val
    return result


def _parse_list_field(raw: str) -> list[str]:
    """解析「a、b、c」/「a,b,c」/「a\nb\nc」格式为 list。"""
    if not raw:
        return []
    parts = re.split(r"[、,\n]", raw)
    return [p.strip() for p in parts if p.strip()]


def _parse_int_or_none(raw: str) -> int | None:
    if not raw or raw.strip().lower() in {"none", "null", "-", ""}:
        return None
    try:
        return int(float(raw.strip()))
    except (ValueError, TypeError):
        return None


# ── 行解析 ───────────────────────────────────────────────────────────


def _parse_row(
    cells: list[str],
    column_map: dict[str, int] | None,
    regime_header: dict[str, str] | None = None,
) -> RevisionEntry | None:
    """把一行数据解析为 RevisionEntry。v1/v2 路径分流。"""
    # v1 兜底：没有列映射但有 6 列，按位置解析
    if column_map is None:
        if len(cells) < 6:
            return None
        return _parse_row_v1_by_position(cells)

    # 按列映射解析
    table_code = _cell(cells, column_map.get("table_code"))
    field_code = _cell(cells, column_map.get("field_code"))
    row_label = _cell(cells, column_map.get("row_label"))
    column_label = _cell(cells, column_map.get("column_label"))

    # 表头行识别（首列等于"表号"等别名时跳过）
    if _clean_header(table_code) in _HEADER_ALIASES["table_code"] | {_clean_header(v) for v in _HEADER_ALIASES["table_code"]}:
        return None
    if not table_code and not field_code:
        return None

    change_type = _normalise_change_type(_cell(cells, column_map.get("change_type")))
    change_desc = _cell(cells, column_map.get("change_desc"))

    entry = RevisionEntry(
        table_code=table_code,
        section=_cell(cells, column_map.get("section")) or "PART_I",
        row_label=row_label,
        column_label=column_label,
        change_type=change_type,
        change_desc=change_desc,
        # v2 字段
        revision_id=_cell(cells, column_map.get("revision_id")),
        regime_version=_cell(cells, column_map.get("regime_version")),
        field_code=field_code,
        field_name=_cell(cells, column_map.get("field_name")),
        change_dimensions=_parse_list_field(_cell(cells, column_map.get("change_dimensions"))),
        change_summary=_cell(cells, column_map.get("change_summary")),
        before_value=_parse_before_after(_cell(cells, column_map.get("before_value"))),
        after_value=_parse_before_after(_cell(cells, column_map.get("after_value"))),
        source_document_id=_parse_int_or_none(_cell(cells, column_map.get("source_document_id"))),
        regulation_evidence=_cell(cells, column_map.get("regulation_evidence")),
        evidence_source_ref=_cell(cells, column_map.get("evidence_source_ref")),
        affected_cells=_parse_list_field(_cell(cells, column_map.get("affected_cells"))),
        effective_date=_cell(cells, column_map.get("effective_date")),
        confidence_level=_cell(cells, column_map.get("confidence_level")) or "MEDIUM",
        review_status=_cell(cells, column_map.get("review_status")) or "DRAFT",
    )

    # v1 没有 field_name 时退化用 column_label
    if not entry.field_name and column_label:
        entry.field_name = column_label

    # 从 Sheet 1「制度版本」继承 regime_version / effective_date（行内未填时）
    if regime_header:
        if not entry.regime_version:
            entry.regime_version = regime_header.get("regime_version", "")
        if not entry.effective_date:
            entry.effective_date = regime_header.get("effective_date", "")

    return entry


def _parse_row_v1_by_position(cells: list[str]) -> RevisionEntry | None:
    table_code = cells[0].strip()
    section = cells[1].strip() or "PART_I"
    row_label = cells[2].strip()
    column_label = cells[3].strip()
    change_type = _normalise_change_type(cells[4])
    change_desc = cells[5].strip()
    if not table_code or not row_label or not column_label:
        return None
    if _clean_header(table_code) in {_clean_header(v) for v in _HEADER_ALIASES["table_code"]}:
        return None
    return RevisionEntry(
        table_code=table_code,
        section=section,
        row_label=row_label,
        column_label=column_label,
        change_type=change_type,
        change_desc=change_desc,
        field_name=column_label,
    )


# ── 制度版本 Sheet 解析 ─────────────────────────────────────────────


_REGIME_KEY_ALIASES = {
    "regime_version": {"制度版本", "版本", "regime_version"},
    "effective_date": {"生效日期", "生效报送期", "effective_date"},
    "publish_date": {"发文日期", "publish_date"},
    "reporting_system_code": {"报送体系", "体系", "reporting_system_code"},
    "version_name": {"版本名称", "名称", "version_name"},
    "source_zip_name": {"源zip文件", "源文件", "source_zip_name"},
}


def _parse_regime_header(rows: list[list[str]]) -> dict[str, str]:
    """Sheet 1 是「键 | 值」两列布局，解析为 dict。"""
    result: dict[str, str] = {}
    inv_alias: dict[str, str] = {
        _clean_header(alias): field_name
        for field_name, aliases in _REGIME_KEY_ALIASES.items()
        for alias in aliases
    }
    for cells in rows:
        if len(cells) < 2:
            continue
        key = _clean_header(cells[0])
        val = cells[1].strip() if cells[1] else ""
        field_name = inv_alias.get(key)
        if field_name and val:
            result[field_name] = val
    return result


# ── xls / xlsx 主解析路径 ───────────────────────────────────────────


def _pick_revision_sheet_xlsx(wb) -> tuple[Any, dict[str, str] | None]:
    """从 xlsx workbook 中挑出「修订对照表」sheet + 读「制度版本」sheet 头。"""
    revision_sheet = None
    regime_rows: list[list[str]] | None = None
    for sheet_name in wb.sheetnames:
        if _clean_header(sheet_name) == _clean_header("修订对照表"):
            revision_sheet = wb[sheet_name]
        elif _clean_header(sheet_name) == _clean_header("制度版本"):
            ws = wb[sheet_name]
            regime_rows = [
                [str(v).strip() if v is not None else "" for v in row]
                for row in ws.iter_rows(values_only=True)
            ]
    if revision_sheet is None:
        revision_sheet = wb.active
    regime_header = _parse_regime_header(regime_rows) if regime_rows else None
    return revision_sheet, regime_header


def _parse_xls(file_path: Path) -> list[RevisionEntry]:
    import xlrd
    try:
        wb = xlrd.open_workbook(str(file_path))
    except Exception as exc:
        raise ValueError(f"无法打开修订对照表 {file_path.name}: {exc}") from exc

    # 尝试找「修订对照表」sheet
    revision_sheet = wb.sheet_by_index(0)
    regime_header: dict[str, str] | None = None
    for i in range(wb.nsheets):
        name = wb.sheet_by_index(i).name
        if _clean_header(name) == _clean_header("修订对照表"):
            revision_sheet = wb.sheet_by_index(i)
        elif _clean_header(name) == _clean_header("制度版本"):
            rs = wb.sheet_by_index(i)
            regime_rows = [
                [str(rs.cell(r, c).value).strip() for c in range(rs.ncols)]
                for r in range(rs.nrows)
            ]
            regime_header = _parse_regime_header(regime_rows)

    rows = [
        [str(revision_sheet.cell(row_idx, c).value).strip() for c in range(revision_sheet.ncols)]
        for row_idx in range(revision_sheet.nrows)
    ]
    return _parse_rows(rows, regime_header)


def _parse_xlsx(file_path: Path) -> list[RevisionEntry]:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
    except Exception as exc:
        raise ValueError(f"无法打开修订对照表 {file_path.name}: {exc}") from exc

    revision_sheet, regime_header = _pick_revision_sheet_xlsx(wb)
    if revision_sheet is None:
        return []

    rows = [
        [str(v).strip() if v is not None else "" for v in row]
        for row in revision_sheet.iter_rows(values_only=True)
    ]
    return _parse_rows(rows, regime_header)


def _parse_rows(
    rows: list[list[str]],
    regime_header: dict[str, str] | None,
) -> list[RevisionEntry]:
    """共用的行解析逻辑：探测表头 → 逐行解析。"""
    header = _detect_header_map(rows)
    start_row = header[0] + 1 if header else 0
    column_map = header[1] if header else None

    entries: list[RevisionEntry] = []
    for cells in rows[start_row:]:
        entry = _parse_row(cells, column_map, regime_header=regime_header)
        if entry:
            entries.append(entry)
    return entries
