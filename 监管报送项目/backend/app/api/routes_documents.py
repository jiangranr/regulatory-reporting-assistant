import json
import re
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlmodel import Session, select

from app.core.config import get_settings
from app.core.database import get_session
from app.models.db_models import (
    DataFieldCatalog,
    DocumentTaskProfile,
    RegDocument,
    RegReportingItem,
    RegReportingObject,
    RegReportingSection,
    ReportingItemLineage,
)
from app.models.enums import DocumentStatus
from app.models.schemas import (
    DocumentTaskProfileRead,
    ExcelDiffEntryRead,
    InstructionChangeAnalysisRead,
    ItemChangeScanResultRead,
    LaneSignalRead,
    RawDocumentFileRead,
    RawDocumentTextRead,
    RegDocumentRead,
    RevisionCandidateRead,
    RevisionEntryRead,
    TableChangeSignalRead,
    TripletUploadResponse,
)
from app.services.document_parser import parse_text_document
from app.services.instruction_change_analyzer import (
    InstructionChangeAnalysis,
    analyze_instruction_changes,
)
from app.services.instruction_parser import (
    extract_regulatory_metadata,
    parse_instruction_file,
    parsed_document_from_pair,
)
from app.services import document_profiler
from app.services.llm_client import LLMClientError
from app.services import excel_parser as _excel_parser
from app.services.revision_table_parser import parse_revision_table
from app.services.g31_excel_diff import diff_excel_with_db, summarise_diff
from app.services.item_change_scanner import scan_and_annotate, build_change_summary
from app.services.reporting_item_scope import filter_analysis_items

router = APIRouter(prefix="/api/documents", tags=["documents"])


def _apply_regulatory_metadata(document: RegDocument, parsed_text: str) -> None:
    """从解析后的正文抽取监管元数据并写入 document（mutating）。

    抽取失败不阻塞上传——document.metadata_extraction_status 反映结果。
    """
    if not parsed_text:
        document.metadata_extraction_status = "FAILED"
        return
    meta = extract_regulatory_metadata(parsed_text)
    document.document_no = meta.document_no
    document.issuing_authority = meta.issuing_authority
    document.published_at = meta.published_at
    document.effective_date = meta.effective_date
    document.first_report_period = meta.first_report_period
    document.regulatory_intent = meta.regulatory_intent
    document.metadata_extraction_status = meta.status


@router.post("/upload", response_model=RegDocumentRead, status_code=status.HTTP_201_CREATED)
async def upload_document(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
) -> RegDocument:
    content = await file.read()
    parsed = parse_text_document(file.filename or "uploaded.txt", content)

    settings = get_settings()
    settings.upload_dir.mkdir(parents=True, exist_ok=True)
    original_filename = file.filename or "uploaded.txt"
    suffix = Path(original_filename).suffix
    storage_path = settings.upload_dir / f"{Path(original_filename).stem}-{uuid4().hex}{suffix}"
    storage_path.write_bytes(content)
    parse_status = DocumentStatus.PARSED if parsed.text else DocumentStatus.FAILED

    document = RegDocument(
        filename=original_filename,
        content_type=file.content_type or "application/octet-stream",
        storage_path=str(storage_path),
        title=Path(original_filename).stem,
        text_excerpt=parsed.excerpt,
        parsed_text=parsed.text,
        status=parse_status,
        parse_status=parse_status,
        parser=parsed.parser,
        char_count=parsed.char_count,
        paragraph_count=parsed.paragraph_count,
        table_count=parsed.table_count,
        parse_quality=parsed.quality,
        parse_error_message=parsed.error_message,
    )
    _apply_regulatory_metadata(document, parsed.text)
    session.add(document)
    session.commit()
    session.refresh(document)
    return document


@router.post("/upload-reporting-pair", response_model=RegDocumentRead, status_code=status.HTTP_201_CREATED)
async def upload_reporting_pair(
    template_file: UploadFile = File(...),
    instruction_file: UploadFile = File(...),
    session: Session = Depends(get_session),
) -> RegDocument:
    template_content = await template_file.read()
    instruction_content = await instruction_file.read()

    settings = get_settings()
    settings.upload_dir.mkdir(parents=True, exist_ok=True)
    template_filename = template_file.filename or "reporting-template"
    instruction_filename = instruction_file.filename or "reporting-instruction"
    template_path = _save_upload_file(template_filename, template_content)
    instruction_path = _save_upload_file(instruction_filename, instruction_content)

    instruction = parse_instruction_file(instruction_filename, instruction_content)
    parsed = parsed_document_from_pair(template_filename, instruction_filename, instruction)
    parse_status = DocumentStatus.PARSED if parsed.text else DocumentStatus.FAILED

    document = RegDocument(
        filename=f"{template_filename} + {instruction_filename}",
        content_type="application/vnd.reg-reporting.pair",
        storage_path=f"{template_path};{instruction_path}",
        title=f"{Path(template_filename).stem} 与 {Path(instruction_filename).stem}",
        text_excerpt=parsed.excerpt,
        parsed_text=parsed.text,
        status=parse_status,
        parse_status=parse_status,
        parser=parsed.parser,
        char_count=parsed.char_count,
        paragraph_count=parsed.paragraph_count,
        table_count=parsed.table_count,
        parse_quality=parsed.quality,
        parse_error_message=parsed.error_message,
    )
    _apply_regulatory_metadata(document, parsed.text)
    session.add(document)
    session.commit()
    session.refresh(document)
    return document


@router.post(
    "/upload-reporting-triplet",
    response_model=TripletUploadResponse,
    status_code=status.HTTP_201_CREATED,
)
async def upload_reporting_triplet(
    object_code: str = "G31",
    template_file: UploadFile | None = File(default=None),
    instruction_file: UploadFile | None = File(default=None),
    revision_table_file: UploadFile | None = File(default=None),
    session: Session = Depends(get_session),
) -> TripletUploadResponse:
    """
    三文件可选上传接口（G31 表样 / 填报说明 / 修订对照表，均可独立上传）。

    处理逻辑：
    - template_file：解析 Excel 结构 → 与库中版本做 diff → 可选入库
    - instruction_file：提取批注/修订/高亮 → 生成组合文档
    - revision_table_file：解析修订对照表 → 精确匹配指标格
    - 若有至少一个信号源，触发 item_change_scanner 更新 change_status
    """
    if not any([template_file, instruction_file, revision_table_file]):
        raise HTTPException(status_code=422, detail="至少需要上传一个文件")

    settings = get_settings()
    settings.upload_dir.mkdir(parents=True, exist_ok=True)

    error_messages: list[str] = []
    filenames: list[str] = []
    storage_paths: list[str] = []

    # ── 解析各文件 ─────────────────────────────────────────────────────────────
    excel_parse_result = None
    template_filename = ""
    template_parsed = False

    instruction_result = None
    instruction_filename = ""
    instruction_parsed = False

    revision_entries = []
    revision_table_parsed = False
    revision_entry_count = 0

    # 1. 报表表样 Excel
    if template_file:
        template_filename = template_file.filename or "reporting-template.xls"
        template_content = await template_file.read()
        template_path = _save_upload_file(template_filename, template_content)
        storage_paths.append(str(template_path))
        filenames.append(template_filename)
        try:
            excel_parse_result = _excel_parser.parse_excel(template_path, object_code)
            template_parsed = True
        except Exception as exc:
            error_messages.append(f"Excel 表样解析失败：{exc}")

    # 2. 填报说明 Word/PDF
    if instruction_file:
        instruction_filename = instruction_file.filename or "instruction.doc"
        instruction_content = await instruction_file.read()
        instruction_path = _save_upload_file(instruction_filename, instruction_content)
        storage_paths.append(str(instruction_path))
        filenames.append(instruction_filename)
        try:
            instruction_result = parse_instruction_file(instruction_filename, instruction_content)
            instruction_parsed = bool(instruction_result.text)
        except Exception as exc:
            error_messages.append(f"填报说明解析失败：{exc}")

    # 3. 修订对照表 Excel
    if revision_table_file:
        rev_filename = revision_table_file.filename or "revision-table.xlsx"
        rev_content = await revision_table_file.read()
        rev_path = _save_upload_file(rev_filename, rev_content)
        storage_paths.append(str(rev_path))
        filenames.append(rev_filename)
        try:
            revision_entries = parse_revision_table(rev_path)
            revision_table_parsed = True
            revision_entry_count = len(revision_entries)
        except Exception as exc:
            error_messages.append(f"修订对照表解析失败：{exc}")

    # ── Excel 差异检测 ─────────────────────────────────────────────────────────
    excel_diffs = []
    excel_diff_count = 0
    if excel_parse_result:
        try:
            excel_diffs = diff_excel_with_db(session, object_code, excel_parse_result)
            excel_diff_count = len(excel_diffs)
        except Exception as exc:
            error_messages.append(f"Excel 差异检测失败：{exc}")

    # ── 指标变更扫描 ───────────────────────────────────────────────────────────
    scan_results = []
    scan_summary = ""
    revision_lane = []
    excel_lane = []
    instruction_lane = []
    revision_candidates = []
    instruction_text_for_scan = (instruction_result.text if instruction_result else None)
    if revision_entries or excel_diffs or instruction_text_for_scan:
        try:
            scan_report = scan_and_annotate(
                session,
                object_code=object_code,
                revision_entries=revision_entries if revision_entries else None,
                excel_diffs=excel_diffs if excel_diffs else None,
                instruction_text=instruction_text_for_scan,
            )
            scan_results = scan_report.results
            revision_lane = scan_report.revision_lane
            excel_lane = scan_report.excel_lane
            instruction_lane = scan_report.instruction_lane
            revision_candidates = scan_report.revision_candidates
            scan_summary = build_change_summary(scan_results, revision_candidates)
        except Exception as exc:
            error_messages.append(f"指标变更扫描失败：{exc}")

    # ── 构建联合文档文本 ────────────────────────────────────────────────────────
    if instruction_result:
        from app.services.instruction_parser import parsed_document_from_pair
        parsed = parsed_document_from_pair(
            template_filename or "（无表样）",
            instruction_filename,
            instruction_result,
        )
    else:
        from app.models.schemas import ParsedDocument
        combined_parts = [f"报表代码：{object_code}"]
        if template_file:
            combined_parts.append(f"表样文件：{template_filename}")
            if excel_diff_count:
                combined_parts.append(f"Excel 差异：{summarise_diff(excel_diffs)}")
        if revision_table_file:
            combined_parts.append(f"修订对照表：共 {revision_entry_count} 条变更记录")
        if scan_summary:
            combined_parts.append(f"扫描结果：{scan_summary}")
        combined_text = "\n".join(combined_parts)
        parsed = ParsedDocument(
            filename="+".join(filenames) if filenames else object_code,
            text=combined_text,
            excerpt=combined_text[:160],
            parser="triplet",
            char_count=len(combined_text),
            paragraph_count=len(combined_parts),
            quality="GOOD" if not error_messages else "FAILED",
        )

    # ── 保存文档记录 ────────────────────────────────────────────────────────────
    parse_status = DocumentStatus.PARSED if parsed.text else DocumentStatus.FAILED
    document = RegDocument(
        filename=" + ".join(filenames) if filenames else f"{object_code} 三文件",
        content_type="application/vnd.reg-reporting.triplet",
        storage_path=";".join(storage_paths),
        title=f"{object_code} 指标变更扫描",
        text_excerpt=parsed.excerpt,
        parsed_text=parsed.text,
        status=parse_status,
        parse_status=parse_status,
        parser=parsed.parser,
        char_count=parsed.char_count,
        paragraph_count=parsed.paragraph_count,
        table_count=parsed.table_count,
        parse_quality=parsed.quality,
        parse_error_message="\n".join(error_messages),
    )
    _apply_regulatory_metadata(document, parsed.text)
    session.add(document)
    session.commit()
    session.refresh(document)

    doc_read = _normalize_document(document)

    return TripletUploadResponse(
        document=RegDocumentRead.model_validate(doc_read),
        template_parsed=template_parsed,
        instruction_parsed=instruction_parsed,
        revision_table_parsed=revision_table_parsed,
        revision_entry_count=revision_entry_count,
        excel_diff_count=excel_diff_count,
        scan_results=[
            ItemChangeScanResultRead(
                item_code=r.item_code,
                row_label=r.row_label,
                column_label=r.column_label,
                old_status=r.old_status,
                new_status=r.new_status,
                source=r.source,
                confidence=r.confidence,
                evidence=r.evidence,
                overridden_sources=list(r.overridden_sources),
            )
            for r in scan_results
        ],
        revision_lane=[
            LaneSignalRead(
                item_code=s.item_code,
                row_label=s.row_label,
                column_label=s.column_label,
                change_type=s.change_type,
                confidence=s.confidence,
                evidence=s.evidence,
                won=s.won,
            )
            for s in revision_lane
        ],
        excel_lane=[
            LaneSignalRead(
                item_code=s.item_code,
                row_label=s.row_label,
                column_label=s.column_label,
                change_type=s.change_type,
                confidence=s.confidence,
                evidence=s.evidence,
                won=s.won,
            )
            for s in excel_lane
        ],
        instruction_lane=[
            LaneSignalRead(
                item_code=s.item_code,
                row_label=s.row_label,
                column_label=s.column_label,
                change_type=s.change_type,
                confidence=s.confidence,
                evidence=s.evidence,
                won=s.won,
            )
            for s in instruction_lane
        ],
        revision_candidates=[
            RevisionCandidateRead(
                table_code=c.table_code,
                section=c.section,
                row_label=c.row_label,
                column_label=c.column_label,
                change_type=c.change_type,
                change_desc=c.change_desc,
                match_status=c.match_status,
                matched_item_code=c.matched_item_code,
                confidence=c.confidence,
                evidence=c.evidence,
                reason=c.reason,
                # v2 字段维度
                revision_id=c.revision_id,
                regime_version=c.regime_version,
                field_code=c.field_code,
                field_name=c.field_name,
                change_dimensions=c.change_dimensions,
                change_summary=c.change_summary,
                before_value=c.before_value,
                after_value=c.after_value,
                regulation_evidence=c.regulation_evidence,
                evidence_source_ref=c.evidence_source_ref,
                affected_cells=c.affected_cells,
                effective_date=c.effective_date,
                matched_item_codes=c.matched_item_codes,
                composite_match=c.composite_match,
            )
            for c in revision_candidates
        ],
        scan_summary=scan_summary,
        excel_diff=[
            ExcelDiffEntryRead(
                item_code=d.item_code,
                row_label=d.row_label,
                column_label=d.column_label,
                diff_type=d.diff_type,
                old_row_label=d.old_row_label,
                old_col_label=d.old_col_label,
                evidence=d.evidence,
            )
            for d in excel_diffs
        ],
        revision_entries=[
            RevisionEntryRead(
                table_code=e.table_code,
                section=e.section,
                row_label=e.row_label,
                column_label=e.column_label,
                change_type=e.change_type,
                change_desc=e.change_desc,
            )
            for e in revision_entries
        ],
        error_messages=error_messages,
    )


@router.get("", response_model=list[RegDocumentRead])
def list_documents(session: Session = Depends(get_session)) -> list[RegDocument]:
    documents = list(session.exec(select(RegDocument).order_by(RegDocument.created_at.desc())).all())
    return [_normalize_document(document) for document in documents]


@router.get("/{document_id}", response_model=RegDocumentRead)
def get_document(document_id: int, session: Session = Depends(get_session)) -> RegDocument:
    document = session.get(RegDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document not found")
    return _normalize_document(document)


@router.delete("/{document_id}", status_code=status.HTTP_200_OK)
def delete_document(document_id: int, session: Session = Depends(get_session)) -> dict:
    """删除发文及其级联数据：profiles / clauses / tasks / candidates / impacts / tickets。

    返回各类被删除行数，方便前端 toast 提示。
    """
    # Lazy import to avoid top-level circular references
    from app.models.db_models import (
        RegClause,
        RegReportingChangeCandidate,
        RegReportingImpactItem,
        RegSemanticItem,
        RegTask,
        TicketDraft,
    )

    document = session.get(RegDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document not found")

    counts = {"profiles": 0, "clauses": 0, "semantic_items": 0, "tasks": 0,
              "candidates": 0, "impacts": 0, "tickets": 0, "files": 0}

    # 1. profile（直接 document_id 关联）
    for row in session.exec(select(DocumentTaskProfile).where(DocumentTaskProfile.document_id == document_id)).all():
        session.delete(row); counts["profiles"] += 1

    # 2. clauses（含其下 semantic items）
    clause_ids = [c.id for c in session.exec(select(RegClause).where(RegClause.document_id == document_id)).all()]
    if clause_ids:
        for s_item in session.exec(select(RegSemanticItem).where(RegSemanticItem.clause_id.in_(clause_ids))).all():
            session.delete(s_item); counts["semantic_items"] += 1
        for c in session.exec(select(RegClause).where(RegClause.document_id == document_id)).all():
            session.delete(c); counts["clauses"] += 1

    # 3. tasks（含其下 candidates / impacts / tickets）
    task_ids = [t.id for t in session.exec(select(RegTask).where(RegTask.document_id == document_id)).all()]
    if task_ids:
        for cand in session.exec(select(RegReportingChangeCandidate).where(RegReportingChangeCandidate.task_id.in_(task_ids))).all():
            session.delete(cand); counts["candidates"] += 1
        for imp in session.exec(select(RegReportingImpactItem).where(RegReportingImpactItem.task_id.in_(task_ids))).all():
            session.delete(imp); counts["impacts"] += 1
        for tk in session.exec(select(TicketDraft).where(TicketDraft.task_id.in_(task_ids))).all():
            session.delete(tk); counts["tickets"] += 1
        for t in session.exec(select(RegTask).where(RegTask.document_id == document_id)).all():
            session.delete(t); counts["tasks"] += 1

    # 4. 物理文件（多文件用 ; 分隔）
    if document.storage_path:
        for p in document.storage_path.split(";"):
            p = p.strip()
            if not p:
                continue
            try:
                fp = Path(p)
                if fp.exists() and fp.is_file():
                    fp.unlink()
                    counts["files"] += 1
            except OSError:
                pass  # 文件可能已被人为删除，不阻塞主流程

    # 5. 主表
    session.delete(document)
    session.commit()

    return {"document_id": document_id, "deleted": counts}


@router.get("/{document_id}/raw-text", response_model=RawDocumentTextRead)
def get_document_raw_text(document_id: int, session: Session = Depends(get_session)) -> RawDocumentTextRead:
    document = session.get(RegDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document not found")

    normalized = _normalize_document(document)
    files = _collect_raw_document_files(normalized)
    if not files:
        fallback_text = normalized.parsed_text or normalized.text_excerpt or "暂无解析文本。"
        fallback_source = (
            "parsed_text" if normalized.parsed_text
            else ("text_excerpt" if normalized.text_excerpt else "empty")
        )
        files = [
            RawDocumentFileRead(
                filename=normalized.filename or normalized.title,
                role="summary",
                role_label="解析摘要",
                source=fallback_source,
                text=fallback_text,
            )
        ]

    top_source = files[0].source if len(files) == 1 else "multi_file"
    top_text = "\n\n".join(f.text for f in files if f.text) or "暂无解析文本。"
    return RawDocumentTextRead(
        document_id=document_id,
        title=normalized.title,
        filename=normalized.filename,
        source=top_source,
        text=top_text,
        files=files,
    )


@router.post(
    "/{document_id}/profile",
    response_model=DocumentTaskProfileRead,
    status_code=status.HTTP_201_CREATED,
)
def profile_document(document_id: int, session: Session = Depends(get_session)) -> DocumentTaskProfileRead:
    document = session.get(RegDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document not found")

    draft = _triplet_document_profile(_normalize_document(document), session)
    if draft is None:
        try:
            draft = document_profiler.generate_document_profile(
                _normalize_document(document),
                _load_profile_context(session),
                session=session,  # 传 session 让 profiler 用 item_resolver 补落格
            )
        except LLMClientError as exc:
            raise HTTPException(status_code=502, detail=str(exc)) from exc

    # 汇总 change_signals 中的证据文本
    evidence_parts = [s.evidence_text for s in draft.change_signals if s.evidence_text]
    evidence_text = "；".join(evidence_parts[:3])

    profile = DocumentTaskProfile(
        document_id=document_id,
        affected_table_codes=_dump_json(draft.affected_table_codes),
        change_signals=_dump_json([s.model_dump() for s in draft.change_signals]),
        in_scope_tables=_dump_json(draft.in_scope_tables),
        out_of_scope_tables=_dump_json(draft.out_of_scope_tables),
        suggested_route=draft.suggested_route,
        should_create_task=draft.should_create_task,
        confidence_score=draft.confidence_score,
        reason=draft.reason,
        evidence_text=evidence_text,
        llm_model=draft.llm_model,
        raw_response=draft.raw_response,
    )
    session.add(profile)
    session.commit()
    session.refresh(profile)
    return _profile_read(profile)


def _triplet_document_profile(
    document: RegDocument,
    session: Session,
) -> document_profiler.Document1104ProfileDraft | None:
    if document.content_type != "application/vnd.reg-reporting.triplet":
        return None

    storage_paths = _triplet_storage_paths(document)
    revision_paths = [
        Path(path)
        for path in storage_paths
        if _is_revision_table_path(path)
    ]
    if revision_paths:
        return _triplet_revision_document_profile(document, session, revision_paths)

    template_paths = [
        Path(path)
        for path in storage_paths
        if _is_excel_template_path(path)
    ]
    if template_paths:
        return _triplet_excel_document_profile(document, session, template_paths)

    return None


def _triplet_revision_document_profile(
    document: RegDocument,
    session: Session,
    revision_paths: list[Path],
) -> document_profiler.Document1104ProfileDraft | None:
    del document

    revision_entries = []
    for path in revision_paths:
        if not path.exists():
            continue
        revision_entries.extend(parse_revision_table(path))
    if not revision_entries:
        return None

    signals: list[document_profiler.TableChangeSignal] = []
    for table_code in sorted({entry.table_code.upper() for entry in revision_entries if entry.table_code}):
        report = scan_and_annotate(
            session,
            object_code=table_code,
            revision_entries=revision_entries,
        )
        for candidate in report.revision_candidates:
            if candidate.change_type == "UNCHANGED":
                continue
            signals.append(
                document_profiler.TableChangeSignal(
                    table_code=candidate.table_code,
                    section_hint=candidate.section,
                    indicator_hint=_revision_display_name(candidate),
                    change_type=_revision_change_type_to_profile_type(candidate.change_type),
                    evidence_text=_revision_change_explanation(candidate),
                    confidence=candidate.confidence or 0.7,
                    evidence_verified=True,
                    matched_item_code=candidate.matched_item_code,
                    match_status=candidate.match_status,
                    composite_match=candidate.composite_match,
                )
            )

    affected_codes = sorted({signal.table_code for signal in signals})
    in_scope = sorted(code for code in affected_codes if _is_catalog_object_code(session, code))
    out_of_scope = sorted(code for code in affected_codes if code not in in_scope)
    route = document_profiler.FULL_ANALYSIS if signals else document_profiler.MANUAL_REVIEW
    raw_response = json.dumps(
        {
            "source": "triplet_revision_table",
            "revision_entry_count": len(revision_entries),
            "change_signal_count": len(signals),
        },
        ensure_ascii=False,
    )

    return document_profiler.Document1104ProfileDraft(
        has_1104_reference=bool(affected_codes),
        affected_table_codes=affected_codes,
        in_scope_tables=in_scope,
        out_of_scope_tables=out_of_scope,
        change_signals=signals,
        suggested_route=route,
        should_create_task=bool(signals),
        confidence_score=0.9 if signals else 0.5,
        reason=(
            f"基于三文件上传中的修订对照表识别 {len(signals)} 条新增/修改/删除变更信号；"
            "未落格信号仍保留为待确认变更信号。"
        ),
        llm_model="rule-based-triplet-scanner",
        raw_response=raw_response,
    )


def _triplet_excel_document_profile(
    document: RegDocument,
    session: Session,
    template_paths: list[Path],
) -> document_profiler.Document1104ProfileDraft | None:
    object_code = _infer_triplet_object_code(document, template_paths)
    if not object_code:
        return None

    parsed_files: list[str] = []
    excel_diffs = []
    for path in template_paths:
        if not path.exists():
            continue
        parse_result = _excel_parser.parse_excel(path, object_code)
        diffs = diff_excel_with_db(session, object_code, parse_result)
        if diffs:
            excel_diffs.extend(diffs)
        parsed_files.append(_original_filename_from_storage_path(path))

    if not parsed_files:
        return None

    if excel_diffs:
        scan_and_annotate(session, object_code=object_code, excel_diffs=excel_diffs)

    signals = _excel_diffs_to_profile_signals(object_code, excel_diffs)
    affected_codes = [object_code] if object_code else sorted({signal.table_code for signal in signals})
    in_scope = sorted(code for code in affected_codes if _is_catalog_object_code(session, code))
    out_of_scope = sorted(code for code in affected_codes if code not in in_scope)
    route = document_profiler.FULL_ANALYSIS if signals else document_profiler.MANUAL_REVIEW
    raw_response = json.dumps(
        {
            "source": "triplet_excel_diff",
            "object_code": object_code,
            "template_files": parsed_files,
            "excel_diff_count": len(excel_diffs),
            "change_signal_count": len(signals),
            "diff_counts": _excel_diff_counts(excel_diffs),
        },
        ensure_ascii=False,
    )

    return document_profiler.Document1104ProfileDraft(
        has_1104_reference=bool(affected_codes),
        affected_table_codes=affected_codes,
        in_scope_tables=in_scope,
        out_of_scope_tables=out_of_scope,
        change_signals=signals,
        suggested_route=route,
        should_create_task=bool(signals),
        confidence_score=0.82 if signals else 0.7,
        reason=(
            f"基于三文件上传中的新版 Excel 表样与系统基线目录比对，识别 {len(signals)} 条"
            "新增/删除/标签变化信号；该路径未使用 LLM 生成变更事实。"
            if signals
            else "新版 Excel 表样已完成结构化比对，未发现新增/删除/标签变化信号。"
        ),
        llm_model="rule-based-triplet-excel-diff",
        raw_response=raw_response,
    )


def _excel_diffs_to_profile_signals(
    object_code: str,
    excel_diffs: list[Any],
) -> list[document_profiler.TableChangeSignal]:
    signals: list[document_profiler.TableChangeSignal] = []
    seen: set[tuple[str, str, str, str]] = set()
    for diff in excel_diffs:
        change_type = _excel_diff_type_to_profile_type(diff.diff_type)
        if change_type == "UNCLEAR":
            continue
        indicator_hint = _excel_diff_indicator_hint(diff)
        key = (change_type, diff.item_code, diff.row_label, diff.column_label)
        if key in seen:
            continue
        seen.add(key)
        match_status = {
            "NEW": "UNMATCHED_NEW",
            "REMOVED": "MATCHED_EXISTING",
            "LABEL_CHANGED": "MATCHED_EXISTING",
        }.get(diff.diff_type, "")
        matched_item_code = diff.item_code if diff.diff_type in {"REMOVED", "LABEL_CHANGED"} else ""
        composite_match = {}
        if diff.diff_type == "NEW":
            composite_match = {
                "match_type": "EXCEL_NEW_ITEM",
                "proposed_item_code": diff.item_code,
                "row_label": diff.row_label,
                "column_label": diff.column_label,
            }
        signals.append(
            document_profiler.TableChangeSignal(
                table_code=object_code,
                section_hint=_section_hint_from_item_code(diff.item_code),
                indicator_hint=indicator_hint,
                change_type=change_type,
                evidence_text=diff.evidence,
                confidence=0.8 if diff.diff_type != "NEW" else 0.75,
                evidence_verified=True,
                matched_item_code=matched_item_code,
                match_status=match_status,
                composite_match=composite_match,
            )
        )
    return signals


def _excel_diff_type_to_profile_type(diff_type: str) -> str:
    return {
        "NEW": "ADD",
        "REMOVED": "DELETE",
        "LABEL_CHANGED": "MODIFY",
    }.get(diff_type, "UNCLEAR")


def _excel_diff_indicator_hint(diff: Any) -> str:
    row = str(getattr(diff, "row_label", "") or "").strip()
    col = str(getattr(diff, "column_label", "") or "").strip()
    if row and col:
        return f"{row} × {col}"
    return row or col or str(getattr(diff, "item_code", "") or "").strip()


def _section_hint_from_item_code(item_code: str) -> str:
    parts = (item_code or "").split(".")
    return parts[1] if len(parts) > 1 else ""


def _excel_diff_counts(excel_diffs: list[Any]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for diff in excel_diffs:
        counts[diff.diff_type] = counts.get(diff.diff_type, 0) + 1
    return counts


def _triplet_storage_paths(document: RegDocument) -> list[Path]:
    return [
        Path(path.strip())
        for path in (document.storage_path or "").split(";")
        if path and path.strip()
    ]


def _is_revision_table_path(path: Path) -> bool:
    # 修订对照表只能是 Excel 格式；纯文本文件即便文件名含 revision 也不是对照表
    if path.suffix.lower() not in {".xls", ".xlsx"}:
        return False
    lower = path.name.lower()
    return "修订" in path.name or "revision" in lower


def _is_excel_template_path(path: Path) -> bool:
    return path.suffix.lower() in {".xls", ".xlsx"} and not _is_revision_table_path(path)


_TRIPLET_OBJECT_CODE_RE = re.compile(r"\b([A-Z]\d{1,2}[A-Z]?)\b")


def _infer_triplet_object_code(document: RegDocument, paths: list[Path]) -> str:
    candidates = [
        document.title or "",
        document.filename or "",
        document.parsed_text or "",
        document.text_excerpt or "",
    ]
    candidates.extend(path.name for path in paths)
    for text in candidates:
        match = _TRIPLET_OBJECT_CODE_RE.search(text.upper())
        if match:
            return match.group(1)
    return ""


def _revision_change_type_to_profile_type(change_type: str) -> str:
    return {
        "NEW": "ADD",
        "MODIFIED": "MODIFY",
        "DELETED": "DELETE",
    }.get(change_type, "UNCLEAR")


def _is_catalog_object_code(session: Session, code: str) -> bool:
    return session.exec(
        select(RegReportingObject).where(RegReportingObject.object_code == code)
    ).first() is not None


def _build_document_raw_text(document: RegDocument) -> tuple[str, str]:
    if document.content_type == "application/vnd.reg-reporting.triplet":
        revision_text = _build_triplet_revision_raw_text(document)
        if revision_text:
            return "revision_table", revision_text

    if document.parsed_text:
        return "parsed_text", document.parsed_text
    if document.text_excerpt:
        return "text_excerpt", document.text_excerpt
    return "empty", "暂无解析文本。"


_RAW_TEXT_ROLE_LABELS = {
    "template": "新表样",
    "instruction": "填报说明",
    "revision_table": "修订对照表",
    "summary": "解析摘要",
    "unknown": "其他",
}

_RAW_TEXT_MAX_CHARS = 80000
_RAW_TEXT_MAX_EXCEL_ROWS = 400
_UPLOAD_UUID_SUFFIX_LEN = 32
_UUID_HEX_CHARS = set("0123456789abcdef")


def _original_filename_from_storage_path(path: Path) -> str:
    """`G31(252)-<uuid32hex>.xls` → `G31(252).xls`。无 uuid 后缀则原样返回。"""
    stem = path.stem
    if (
        len(stem) > _UPLOAD_UUID_SUFFIX_LEN + 1
        and stem[-(_UPLOAD_UUID_SUFFIX_LEN + 1)] == "-"
        and all(ch in _UUID_HEX_CHARS for ch in stem[-_UPLOAD_UUID_SUFFIX_LEN:])
    ):
        return stem[: -(_UPLOAD_UUID_SUFFIX_LEN + 1)] + path.suffix
    return path.name


def _infer_raw_file_role(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    lower = filename.lower()
    # 修订对照表只可能是 Excel 文件；txt/pdf 等普通发文即便文件名含 revision 也不是对照表
    if suffix in {".xls", ".xlsx"} and ("修订" in filename or "revision" in lower):
        return "revision_table"
    if suffix in {".doc", ".docx"}:
        return "instruction"
    if suffix in {".xls", ".xlsx"}:
        return "template"
    return "unknown"


def _truncate_raw_text(text: str) -> str:
    if len(text) <= _RAW_TEXT_MAX_CHARS:
        return text
    return text[:_RAW_TEXT_MAX_CHARS] + f"\n\n……（文本超过 {_RAW_TEXT_MAX_CHARS} 字符，已截断）"


def _render_xls_text(path: Path) -> str:
    import xlrd
    workbook = xlrd.open_workbook(str(path))
    lines: list[str] = []
    for sheet_idx in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_idx)
        lines.append(f"# Sheet：{sheet.name}（{sheet.nrows} 行 × {sheet.ncols} 列）")
        row_cap = min(sheet.nrows, _RAW_TEXT_MAX_EXCEL_ROWS)
        for r in range(row_cap):
            cells = []
            for c in range(sheet.ncols):
                value = sheet.cell(r, c).value
                cells.append("" if value in ("", None) else str(value).strip())
            lines.append("\t".join(cells).rstrip())
        if sheet.nrows > row_cap:
            lines.append(f"……（共 {sheet.nrows} 行，仅展示前 {row_cap} 行）")
        lines.append("")
    return "\n".join(lines).rstrip()


def _render_xlsx_text(path: Path) -> str:
    from openpyxl import load_workbook
    workbook = load_workbook(str(path), data_only=True, read_only=True)
    lines: list[str] = []
    try:
        for sheet in workbook.worksheets:
            lines.append(f"# Sheet：{sheet.title}")
            row_count = 0
            truncated = False
            for row in sheet.iter_rows(values_only=True):
                if row_count >= _RAW_TEXT_MAX_EXCEL_ROWS:
                    truncated = True
                    break
                cells = ["" if v is None else str(v).strip() for v in row]
                lines.append("\t".join(cells).rstrip())
                row_count += 1
            if truncated:
                lines.append(f"……（仅展示前 {_RAW_TEXT_MAX_EXCEL_ROWS} 行）")
            lines.append("")
    finally:
        workbook.close()
    return "\n".join(lines).rstrip()


def _render_excel_text(path: Path) -> str:
    if path.suffix.lower() == ".xlsx":
        return _render_xlsx_text(path)
    return _render_xls_text(path)


def _format_revision_entries_text(entries) -> str:
    if not entries:
        return ""
    lines = [
        f"修订对照表解析明细：共 {len(entries)} 条",
        "",
        "序号 | 报表 | 分区 | 指标 | 字段 | 类型 | 调整说明",
    ]
    for idx, entry in enumerate(entries, start=1):
        lines.append(
            " | ".join(
                [
                    str(idx),
                    entry.table_code or "-",
                    entry.section or "-",
                    _revision_display_name(entry) or "-",
                    _revision_field_label(entry) or "-",
                    entry.change_type or "-",
                    _revision_change_explanation(entry) or "-",
                ]
            )
        )
    return "\n".join(lines)


def _build_raw_file_entry(path: Path) -> RawDocumentFileRead:
    original_filename = _original_filename_from_storage_path(path)
    role = _infer_raw_file_role(original_filename)
    role_label = _RAW_TEXT_ROLE_LABELS.get(role, "其他")

    if not path.exists():
        return RawDocumentFileRead(
            filename=original_filename,
            role=role,
            role_label=role_label,
            source=role,
            text="",
            error_message=f"文件不存在：{path.name}",
        )

    try:
        if role == "revision_table":
            entries = parse_revision_table(path)
            text = _format_revision_entries_text(entries) or "暂无修订条目。"
            source = "revision_table"
        elif role == "instruction":
            content = path.read_bytes()
            result = parse_instruction_file(original_filename, content)
            text = result.text or "（未提取到文本）"
            source = result.parser or "instruction"
        elif role == "template":
            text = _render_excel_text(path) or "（Excel 为空）"
            source = "excel_template"
        else:
            content = path.read_bytes()
            parsed = parse_text_document(original_filename, content)
            text = parsed.text or "（未提取到文本）"
            source = parsed.parser or "text"
    except Exception as exc:
        return RawDocumentFileRead(
            filename=original_filename,
            role=role,
            role_label=role_label,
            source=role,
            text="",
            error_message=f"解析失败：{exc}",
        )

    return RawDocumentFileRead(
        filename=original_filename,
        role=role,
        role_label=role_label,
        source=source,
        text=_truncate_raw_text(text),
    )


def _collect_raw_document_files(document: RegDocument) -> list[RawDocumentFileRead]:
    paths = [Path(p) for p in (document.storage_path or "").split(";") if p]
    if not paths:
        return []
    return [_build_raw_file_entry(p) for p in paths]


def _build_triplet_revision_raw_text(document: RegDocument) -> str:
    paths = [
        Path(path)
        for path in (document.storage_path or "").split(";")
        if path and Path(path).suffix.lower() in {".xls", ".xlsx"}
        and ("修订" in Path(path).name or "revision" in Path(path).name.lower())
    ]
    entries = []
    for path in paths:
        if path.exists():
            entries.extend(parse_revision_table(path))
    if not entries:
        return ""

    lines = [
        f"修订对照表解析明细：共 {len(entries)} 条",
        "",
        "序号 | 报表 | 分区 | 指标 | 字段 | 类型 | 调整说明",
    ]
    for index, entry in enumerate(entries, start=1):
        lines.append(
            " | ".join(
                [
                    str(index),
                    entry.table_code or "-",
                    entry.section or "-",
                    _revision_display_name(entry) or "-",
                    _revision_field_label(entry) or "-",
                    entry.change_type or "-",
                    _revision_change_explanation(entry) or "-",
                ]
            )
        )
    return "\n".join(lines)


def _revision_display_name(item: Any) -> str:
    """Return the business-friendly display name for v1/v2 revision entries."""
    row_label = str(getattr(item, "row_label", "") or "").strip()
    column_label = str(getattr(item, "column_label", "") or "").strip()
    field_name = str(getattr(item, "field_name", "") or "").strip()
    field_code = str(getattr(item, "field_code", "") or "").strip()
    change_summary = str(getattr(item, "change_summary", "") or "").strip()

    if row_label and column_label:
        return f"{row_label} × {column_label}"
    if field_name:
        return field_name
    if row_label or column_label:
        return row_label or column_label
    return field_code or change_summary


def _revision_field_label(item: Any) -> str:
    field_code = str(getattr(item, "field_code", "") or "").strip()
    column_label = str(getattr(item, "column_label", "") or "").strip()
    field_name = str(getattr(item, "field_name", "") or "").strip()
    return column_label or field_code or field_name


def _revision_change_explanation(item: Any) -> str:
    change_desc = str(getattr(item, "change_desc", "") or "").strip()
    change_summary = str(getattr(item, "change_summary", "") or "").strip()
    evidence = str(getattr(item, "evidence", "") or "").strip()
    regulation_evidence = str(getattr(item, "regulation_evidence", "") or "").strip()
    return change_desc or change_summary or evidence or regulation_evidence


@router.get("/{document_id}/profile", response_model=DocumentTaskProfileRead | None)
def get_document_profile(
    document_id: int, session: Session = Depends(get_session)
) -> DocumentTaskProfileRead | None:
    profile = session.exec(
        select(DocumentTaskProfile)
        .where(DocumentTaskProfile.document_id == document_id)
        .order_by(DocumentTaskProfile.created_at.desc())
    ).first()
    if profile is None:
        return None
    return _profile_read(profile)


def _normalize_document(document: RegDocument) -> RegDocument:
    document.parse_error_message = document.parse_error_message or ""
    document.parser = document.parser or ""
    document.parse_quality = document.parse_quality or "UNKNOWN"
    document.parse_status = document.parse_status or document.status
    document.char_count = document.char_count or 0
    document.paragraph_count = document.paragraph_count or 0
    document.table_count = document.table_count or 0
    return document


def _save_upload_file(filename: str, content: bytes) -> Path:
    settings = get_settings()
    settings.upload_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(filename).suffix
    storage_path = settings.upload_dir / f"{Path(filename).stem}-{uuid4().hex}{suffix}"
    storage_path.write_bytes(content)
    return storage_path


def _load_profile_context(session: Session) -> dict:
    reporting_objects = session.exec(select(RegReportingObject)).all()
    reporting_sections = session.exec(select(RegReportingSection)).all()
    reporting_items = filter_analysis_items(session.exec(select(RegReportingItem)).all())
    data_fields = session.exec(select(DataFieldCatalog)).all()
    lineage = session.exec(select(ReportingItemLineage)).all()
    field_by_id = {item.id: item for item in data_fields}
    item_by_id = {item.id: item for item in reporting_items}
    object_by_id = {item.id: item for item in reporting_objects}
    return {
        "reporting_objects": [
            {
                "object_code": item.object_code,
                "object_name": item.object_name,
                "object_category": item.object_category,
                "report_frequency": item.report_frequency,
            }
            for item in reporting_objects
        ],
        "reporting_sections": [
            {
                "object_code": object_by_id[item.reporting_object_id].object_code,
                "section_code": item.section_code,
                "section_name": item.section_name,
                "display_order": item.display_order,
                "description": item.description,
            }
            for item in reporting_sections
            if item.reporting_object_id in object_by_id
        ],
        "reporting_items": [
            {
                "item_code": item.item_code,
                "item_name": item.item_name,
                "row_label": item.row_label,
                "column_label": item.column_label,
                "measure_type": item.measure_type,
            }
            for item in reporting_items
        ],
        "data_fields": [
            {
                "field_code": item.field_code,
                "field_name": item.field_name,
                "table_name": item.table_name,
                "column_name": item.column_name,
                "business_meaning": item.business_meaning,
                "owner_team": item.owner_team,
            }
            for item in data_fields
        ],
        "reporting_lineage": [
            {
                "reporting_item_code": item_by_id[item.reporting_item_id].item_code,
                "data_field_code": field_by_id[item.data_field_id].field_code,
                "lineage_role": item.lineage_role,
                "confidence_level": item.confidence_level,
            }
            for item in lineage
            if item.reporting_item_id in item_by_id and item.data_field_id in field_by_id
        ],
    }


def _profile_read(profile: DocumentTaskProfile) -> DocumentTaskProfileRead:
    return DocumentTaskProfileRead(
        id=profile.id,
        document_id=profile.document_id,
        document_type=profile.document_type or "",
        affected_table_codes=_load_json_list(profile.affected_table_codes),
        change_signals=_load_change_signals(profile.change_signals),
        in_scope_tables=_load_json_list(profile.in_scope_tables),
        out_of_scope_tables=_load_json_list(profile.out_of_scope_tables),
        suggested_route=profile.suggested_route,
        should_create_task=profile.should_create_task,
        confidence_score=profile.confidence_score,
        reason=profile.reason or "",
        evidence_text=profile.evidence_text or "",
        llm_model=profile.llm_model or "",
        review_status=profile.review_status or "PENDING",
        raw_response=profile.raw_response or "",
        created_at=profile.created_at,
    )


def _load_change_signals(value: str) -> list[TableChangeSignalRead]:
    try:
        raw = json.loads(value or "[]")
    except json.JSONDecodeError:
        return []
    if not isinstance(raw, list):
        return []
    signals = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        try:
            signals.append(
                TableChangeSignalRead(
                    table_code=item.get("table_code", ""),
                    section_hint=item.get("section_hint", ""),
                    indicator_hint=item.get("indicator_hint", ""),
                    change_type=item.get("change_type", "UNCLEAR"),
                    evidence_text=item.get("evidence_text", ""),
                    confidence=float(item.get("confidence", 0.7)),
                    evidence_verified=bool(item.get("evidence_verified", True)),
                    matched_item_code=item.get("matched_item_code", ""),
                    match_status=item.get("match_status", ""),
                    composite_match=item.get("composite_match") if isinstance(item.get("composite_match"), dict) else {},
                )
            )
        except Exception:
            continue
    return signals


def _dump_json(value: list[str]) -> str:
    return json.dumps(value, ensure_ascii=False)


def _load_json_list(value: str) -> list[str]:
    try:
        loaded = json.loads(value or "[]")
    except json.JSONDecodeError:
        return []
    return loaded if isinstance(loaded, list) else []


@router.post(
    "/{document_id}/analyze-instruction-changes",
    response_model=InstructionChangeAnalysisRead,
)
async def analyze_document_instruction_changes(
    document_id: int,
    object_code: str = "G31",
    session: Session = Depends(get_session),
) -> InstructionChangeAnalysisRead:
    """
    综合 Word 批注、修订追踪、彩色高亮与库中报表范围，
    用 LLM 输出本版本填报说明相对上一版本的变更要点。

    对于联合上传文档（content_type=application/vnd.reg-reporting.pair），
    自动识别填报说明文件路径；单文件上传则直接解析该文件。
    """
    document = session.get(RegDocument, document_id)
    if document is None:
        raise HTTPException(status_code=404, detail="Document not found")

    # 定位填报说明文件
    storage_path = document.storage_path or ""
    paths = [p.strip() for p in storage_path.split(";") if p.strip()]
    if not paths:
        raise HTTPException(status_code=422, detail="文档没有有效的存储路径")

    # 联合上传：第二段是填报说明；单文件：直接用自身
    instruction_path = Path(paths[1] if len(paths) > 1 else paths[0])
    if not instruction_path.exists():
        raise HTTPException(status_code=422, detail=f"填报说明文件不存在：{instruction_path.name}")

    suffix = instruction_path.suffix.lower()
    if suffix not in {".doc", ".docx", ".pdf", ".txt", ".md"}:
        raise HTTPException(status_code=422, detail=f"不支持的填报说明格式：{suffix}")

    instruction_content = instruction_path.read_bytes()
    instruction = parse_instruction_file(instruction_path.name, instruction_content)

    context = _load_profile_context(session)

    analysis: InstructionChangeAnalysis = await analyze_instruction_changes(
        instruction=instruction,
        object_code=object_code,
        reporting_objects=context["reporting_objects"],
        reporting_sections=context["reporting_sections"],
        reporting_items=context["reporting_items"],
    )

    if analysis.error_message:
        raise HTTPException(status_code=502, detail=analysis.error_message)

    from app.models.schemas import ChangeItemRead
    return InstructionChangeAnalysisRead(
        object_code=analysis.object_code,
        summary=analysis.summary,
        key_changes=[
            ChangeItemRead(category=c.category, change=c.change, evidence=c.evidence)
            for c in analysis.key_changes
        ],
        attention_items=analysis.attention_items,
        uncertain_items=analysis.uncertain_items,
        comments_count=analysis.comments_count,
        revisions_count=analysis.revisions_count,
        highlights_count=analysis.highlights_count,
        llm_model=analysis.llm_model,
    )
